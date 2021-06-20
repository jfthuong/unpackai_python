import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Union

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet  # for type hints

# cspell:word modif
# cspell:ignore dateutil, relativedelta


PathLike = Union[str, Path]


@dataclass
class FileInfo:
    path: Path

    @property
    def size(self) -> int:
        return self.path.stat().st_size

    @property
    def friendly_size(self) -> str:
        """Convert a size in bytes (as float) to a size with unit (as a string)"""
        size = float(self.size)
        unit = "B"
        # Reminder: 1 KB = 1024 B, and 1 MB = 1024 KB, ...
        for letter in "KMG":
            if size > 1024:
                size /= 1024
                unit = f"{letter}B"

        # We want to keep 2 digits after floating point
        # because it is a good balance and precision and concision
        return f"{size:0.2f} {unit}"

    @property
    def modif_time(self) -> datetime:
        return datetime.fromtimestamp(self.path.stat().st_mtime)


def _write_list_files(ws: Worksheet, rootdir: Path):
    """Write the list of files in the given worksheet"""
    headers = [
        ("File", 100),
        ("Type", 10),
        ("Size", 15),
        ("Size (bytes)", 15),
        ("Last Modif Time", 20),
    ]
    for col, (head, width) in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = head
        ws.column_dimensions[get_column_letter(col)].width = width

    n_row = 1
    for path in rootdir.rglob("*.*"):
        if not path.is_file():
            continue

        n_row += 1
        info = FileInfo(path)
        infos = [
            str(info.path),
            info.path.suffix,
            info.friendly_size,
            info.size,
            info.modif_time,
        ]
        for col, value in enumerate(infos, start=1):
            cell = ws.cell(row=n_row, column=col)
            cell.value = value
            if isinstance(value, (float, int)):
                cell.number_format = "#,##0_);(#,##0)"
            elif isinstance(value, datetime):
                cell.number_format = "[$-en-US]m/d/yy h:mm AM/PM;@"

    last_col = get_column_letter(len(headers))
    table = Table(displayName="ListFiles", ref=f"A1:{last_col}{n_row}")
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)


def _write_summary(ws: Worksheet):
    """Create a summary of data in a given Worksheet"""
    # TODO: implement


def files_2_excel(root: PathLike, xlsx: PathLike):
    """List all files in a directory and add to an Excel file"""
    wb = Workbook()

    ws_list = wb.active
    ws_list.title = "List of Files"
    _write_list_files(ws_list, Path(root))

    ws_summary = wb.create_sheet("Summary", 0)
    _write_summary(ws_summary)

    wb.save(xlsx)
    print(f"List of files saved successfully in {xlsx}")


if __name__ == "__main__":
    root = Path(__file__).parent.parent
    xlsx = Path(__file__).with_name("list_files.xlsx")
    files_2_excel(root, xlsx)
    os.startfile(xlsx)
