import os
from datetime import datetime
from pathlib import Path
from typing import List

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet  # for type hints
import pandas as pd


# cspell:word modif, ipynb
# cspell:ignore dateutil, relativedelta, ISNUMBER


KEY_LAST_MODIF = "Last Modif Time"
KEY_SIZE_B = "Size (b)"

TABLE_LIST_FILES = "ListFiles"
FORMAT_DATE = "[$-en-US]m/d/yy h:mm AM/PM;@"


def friendly_size(size: float) -> str:
    """Convert a size in bytes (as float) to a size with unit (as a string)"""
    unit = "B"
    # Reminder: 1 KB = 1024 B, and 1 MB = 1024 KB, ...
    for letter in "KMG":
        if size > 1024:
            size /= 1024
            unit = f"{letter}B"

    # We want to keep 2 digits after floating point
    # because it is a good balance and precision and concision
    return f"{size:0.2f} {unit}"


def get_file_type(path: Path):
    """Get the type of file"""
    # We might not have a suffix for files like ".gitignore"
    if path.name.startswith("."):
        return path.name.lstrip(".")

    types = {
        ".html": "HTML",
        ".htm": "HTML",
        ".ipynb": "Jupyter",
        ".xlsx": "Excel",
        ".xls": "Excel",
        ".docx": "MS Word",
        ".doc": "MS Word",
        ".txt": "Text",
        ".py": "Python",
        ".csv": "Data",
        ".json": "Data",
        ".yaml": "Data",
        ".bat": "Batch",
        ".cmd": "Batch",
        ".sh": "Batch",
    }
    return types.get(path.suffix, path.suffix.lstrip("."))


def iter_files(root: Path, exclude_dir: List[str]):
    """Return all info of files found in a root directory"""
    for f in root.rglob("*.*"):
        if f.is_file() and not any(d in f.parts for d in exclude_dir):
            size = f.stat().st_size
            yield {
                "Name": f.name,
                "Path": f.as_posix(),
                "Extension": f.suffix or f.name,
                "Type": get_file_type(f),
                "Size": friendly_size(size),
                KEY_SIZE_B: size,
                KEY_LAST_MODIF: datetime.fromtimestamp(f.stat().st_mtime),
            }


def files_2_df(root: Path, exclude_folders: List[str] = None) -> pd.DataFrame:
    if exclude_folders is None:
        exclude_folders = []
    return pd.DataFrame(iter_files(root, exclude_folders))


def _excel_write_list_files(ws: Worksheet, df: pd.DataFrame):
    """Write the list of files in the given worksheet"""
    ws.title = "List of Files"

    n_row = 0
    for r in dataframe_to_rows(df, index=False, header=True):
        n_row += 1
        ws.append(r)

    # We need to adjust the column width and styles
    ws.delete_rows(2)
    width_format = {
        "Name": (30, None),
        "Path": (100, None),
        "Extension": (10, None),
        "Type": (10, None),
        "Size": (10, None),
        KEY_SIZE_B: (10, "#,##0_);(#,##0)"),
        KEY_LAST_MODIF: (20, FORMAT_DATE),
    }
    for j in range(1, len(df.columns) + 1):
        header = ws.cell(row=1, column=j).value
        width, nb_format = width_format.get(header, (None, None))
        col = get_column_letter(j)
        if width:
            ws.column_dimensions[col].width = width
        if nb_format:
            for cell in ws[col]:
                cell.number_format = nb_format

    # Add a Table with style with striped rows and banded columns
    last_col = get_column_letter(len(df.columns))
    table = Table(displayName=TABLE_LIST_FILES, ref=f"A1:{last_col}{n_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)


def _write_top_ten(
    ws: Worksheet, df_ext: pd.DataFrame, by: str, init_row: int, col_nb: int, descr: str
):
    """Write top 10 File Types with min / max / nb / sum and return last row"""
    df_top = df_ext.sort_values(by=by, ascending=False).reset_index()[:10]
    ws.append([f"Top 10 Types by {descr}"])
    ws.cell(row=init_row, column=1).style = "Headline 1"

    n_row = init_row
    for r in dataframe_to_rows(df_top, index=False, header=True):
        n_row += 1
        ws.append([None] + r)

    # We want to store the values as a table to have a nice style
    table = Table(
        displayName=f"TopType{by.capitalize()}", ref=f"B{init_row + 1}:F{n_row}"
    )
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # We want to create a Pie Chart with the top element separated
    # Note that we have an empty column so we need to shift
    pie = PieChart()
    labels = Reference(ws, min_col=2, min_row=init_row + 2, max_row=n_row)
    data = Reference(ws, min_col=col_nb + 1, min_row=init_row + 2, max_row=n_row)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = f"Top 10 Types by {descr}"
    pie.dataLabels = DataLabelList(showPercent=True)
    slice_ = DataPoint(idx=0, explosion=20)
    pie.series[0].data_points = [slice_]
    ws.add_chart(pie, f"H{init_row}")


def _write_modif_times(ws: Worksheet, init_row: int):
    """Write a table and put graph based on modification time"""
    ws.append([f"Nb of files by modification time"])
    ws.cell(row=init_row, column=1).style = "Headline 1"

    now = datetime.now()
    modif_times = [
        ("2 hours", relativedelta(hours=2)),
        ("3 days", relativedelta(days=3)),
        ("2 weeks", relativedelta(weeks=2)),
        ("6 months", relativedelta(months=6)),
        ("1 year", relativedelta(years=1)),
        ("3 years", relativedelta(years=3)),
    ]

    # We will start to write in column A to have correct column width
    # Also, we will first create the table because we use Table Formula
    n_row = init_row + 1
    end_row = n_row + len(modif_times) + 1
    table = Table(displayName=f"ModifStat", ref=f"A{n_row}:C{end_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    ws.append(["Time period", "Max date", "Nb of files"])
    for label, delta in modif_times:
        n_row += 1
        formula = (
            f'=COUNTIF({TABLE_LIST_FILES}[Last Modif Time],">="&B{n_row})'
            f"-SUM(C${init_row+1}:C{n_row - 1})"
        )
        ws.append([label, now - delta, formula])
        ws.cell(row=n_row, column=2).number_format = FORMAT_DATE

    formula_beyond = f'=COUNTIF({TABLE_LIST_FILES}[Last Modif Time],"<"&B{n_row})'
    ws.append(["Above 3 years", "-", formula_beyond])
    n_row += 1

    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=init_row + 2, max_row=n_row)
    data = Reference(ws, min_col=3, min_row=init_row + 2, max_row=n_row)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = f"Number of files by modification time"
    pie.dataLabels = DataLabelList(showPercent=True)
    # TODO: add slice for max value
    # slice_ = DataPoint(idx=0, explosion=20)
    # pie.series[0].data_points = [slice_]
    ws.add_chart(pie, f"H{init_row}")


def _excel_write_summary(ws: Worksheet, rootdir: Path, df: pd.DataFrame):
    """Create a summary of data in a given Worksheet"""
    # We want to display information about the root directory
    # and the time of generation
    ws.append(["Root Directory:", rootdir.as_posix()])
    ws.append(["Generation Time:", datetime.now()])
    ws.cell(row=2, column=2).number_format = FORMAT_DATE
    for row in (1, 2):
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(italic=True)
    for col in "AB":
        ws.column_dimensions[col].width = 18
    ws.append([])

    # We will then put some summary per file type, along with some graphs
    n_row = 4
    df_ext = df.groupby("Type")[KEY_SIZE_B].agg([min, max, len, sum])

    _write_top_ten(ws, df_ext, "sum", n_row, 5, "Total Size")
    n_row += 12
    for _ in range(3):
        ws.append([])
        n_row += 1

    _write_top_ten(ws, df_ext, "len", n_row, 4, "Nb of files")
    n_row += 12
    for _ in range(3):
        ws.append([])
        n_row += 1

    # We will finally put a table to compute number of files per modification time
    _write_modif_times(ws, n_row)


def df_files_2_excel(xlsx: Path, df: pd.DataFrame, root: Path, do_open=False):
    """Generate Excel from DataFrame (and open at the end if specified)"""
    wb = Workbook()
    ws = wb.active
    _excel_write_list_files(ws, df)

    ws = wb.create_sheet("Summary", 0)
    _excel_write_summary(ws, root, df)

    wb.save(xlsx)
    if do_open:
        os.startfile(xlsx)


if __name__ == "__main__":
    root = Path(__file__).parent.parent.parent
    xlsx = Path(__file__).with_name("list_files.xlsx")
    df = files_2_df(root, exclude_folders=[".svn", ".git", ".mypy_cache"])
    df_files_2_excel(xlsx, df, root, do_open=True)
